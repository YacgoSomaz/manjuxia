import logging
from fastapi import APIRouter, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from typing import List, Optional
from pydantic import BaseModel
from models.storyboards import (
    StoryboardGenerateRequest,
    SplitScenesRequest,
    GenerateSectionRequest,
    RegenerateSingleSectionRequest,
    StoryboardUpdate,
    StoryboardResponse,
    StoryboardGroupedResponse
)
from services.storyboard_service import StoryboardService, register_generation_task, unregister_generation_task, get_running_task_count, cancel_generation_tasks
from services.novel_service import NovelService
from services import sensitive_word_service
from database.db import get_db
import asyncio

import io
import json
import re
from openpyxl import Workbook, load_workbook
from urllib.parse import quote

router = APIRouter(prefix="/api/storyboards", tags=["storyboards"])

logger = logging.getLogger(__name__)

# 限制常量
MAX_EXPORT_COUNT = 200  # 单次导出最大数量
MAX_IMPORT_ROWS = 500   # 单次导入最大行数
MAX_FILE_SIZE = 50 * 1024 * 1024  # 最大文件大小 50MB


_STATE_TAG_RE = re.compile(r"(姿态|情绪|伤势|朝向关系|持有道具)\[")
_STATE_HEADER_RE = re.compile(r"^场景起始状态\s*[:：]?$")
# v3.61.218: 边界 —— 起始状态块到这些行就结束(空间布局/本节主线/镜号N/状态链元数据)
_STATE_BOUNDARY_RE = re.compile(r"^(空间布局|本节主线|镜号\d+|🎬|📏|🔗)")


def _parse_section_start_state_from_text(text: str) -> Optional[dict]:
    """从用户编辑的可见文本解析「场景起始状态」块 → {角色名: 状态}。

    v3.61.218 容错:用户若把「场景起始状态:」表头删了,只要还有
    「名字 = 姿态[..]·情绪[..]..」这类带状态标签的行,也能识别(否则改了不同步)。
    - 有表头:进入块后块内所有 `名字 = 值` 行都收(保持原行为,完全兼容)。
    - 无表头容错:只收 value 带「姿态[/情绪[/伤势[/朝向关系[/持有道具[」标签的行,
      防止把正文里别的 `X = Y` 误判成状态;遇到镜号/🔗/空间布局等边界即停,
      所以绝不会把「🔗 本节结尾状态」误当起始状态(它在镜号之后,早已 break)。
    """
    if not text:
        return None

    parsed: dict = {}
    in_block = False
    for raw_line in text.splitlines():
        line = raw_line.strip()

        if _STATE_HEADER_RE.match(line):
            in_block = True
            continue
        if _STATE_BOUNDARY_RE.match(line):
            if parsed:            # 已收到状态 → 起始状态块结束
                break
            in_block = False      # 还没收到(尚在节头区)→ 继续往下找
            continue
        if not line or "=" not in line:
            continue

        name, value = line.split("=", 1)
        name = name.strip().rstrip(":：").strip()
        value = value.strip()
        if not name or not value:
            continue

        # 有表头:块内全收;无表头:只收带状态标签的行(容错且防误判)
        if in_block or _STATE_TAG_RE.search(value):
            parsed[name] = value
            in_block = True
    return parsed or None


@router.post("/generate")
async def generate_storyboards(request: StoryboardGenerateRequest):
    """生成分镜（同步等待结果）"""
    logger.info(f"[storyboard-api] /generate 被调用: novel_id={request.novel_id}, template_id={request.template_id}, llm_config_id={request.llm_config_id}, script_id={request.script_id}")
    try:
        # 直接同步调用分镜生成，与剧本转换保持一致
        result = await StoryboardService.generate_storyboards(
            novel_id=request.novel_id,
            template_id=request.template_id,
            llm_config_id=request.llm_config_id,
            script_id=request.script_id
        )
        
        logger.info(f"[storyboard-api] /generate 服务返回: success={result.get('success')}, count={result.get('count')}, message={result.get('message')}")
        
        if not result["success"]:
            raise HTTPException(status_code=400, detail=result["message"])
        
        logger.info(f"[storyboard-api] /generate 即将返回响应")
        return result
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[storyboard-api] /generate 发生未预期异常: {type(e).__name__}: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"生成分镜失败: {str(e)}")


@router.post("/split-scenes")
async def split_scenes(request: SplitScenesRequest):
    """拆分剧本场景"""
    logger.info(f"[storyboard-api] /split-scenes 被调用: novel_id={request.novel_id}, script_id={request.script_id}")
    try:
        # 1. 获取剧本内容
        result = await StoryboardService.get_script_content_for_split_scenes(
            novel_id=request.novel_id,
            script_id=request.script_id
        )

        if not result["success"]:
            raise HTTPException(status_code=400, detail=result["message"])

        script_content = result["content"]

        # 2. 拆分场景
        scenes = StoryboardService.split_scenes_from_script(script_content)

        return {
            "success": True,
            "scenes": scenes,
            "total_scenes": len(scenes),
            "message": f"成功拆分出 {len(scenes)} 个场景"
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[storyboard-api] /split-scenes 发生异常: {type(e).__name__}: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"拆分场景失败: {str(e)}")


@router.post("/generate-section")
async def generate_section_storyboards(request: GenerateSectionRequest):
    """为单个场景异步生成分镜"""
    logger.info(f"[storyboard-api] /generate-section 被调用: novel_id={request.novel_id}, "
          f"section_number={request.section_number}, scene_title={request.scene_title}, "
          f"scene_index={request.scene_index}, style_template_id={request.style_template_id}")
    
    # 【关键修复】先同步删除该场景的旧分镜，避免轮询在异步任务执行前读到旧数据
    # 使用 scene_index 精确删除，避免文本匹配不可靠的问题
    if request.scene_index is not None:
        db = await get_db()
        try:
            cursor = await db.execute(
                "DELETE FROM storyboards WHERE novel_id = ? AND script_id = ? AND scene_index = ?",
                (request.novel_id, request.script_id, request.scene_index)
            )
            deleted_count = cursor.rowcount if hasattr(cursor, 'rowcount') else 0
            # 清理关联的 llm_logs
            await db.execute(
                "DELETE FROM llm_logs WHERE novel_id = ? AND source_id = ? AND source_type = 'storyboard' AND source_scene_index = ?",
                (request.novel_id, request.script_id, request.scene_index)
            )
            logger.info(f"[storyboard-api] 同步清理场景 scene_index={request.scene_index} 的 llm_logs")
            await db.commit()
            logger.info(f"[storyboard-api] 同步删除场景 scene_index={request.scene_index} 的旧分镜: deleted_count={deleted_count}")
        except Exception as e:
            logger.error(f"[storyboard-api] 同步删除旧分镜失败: {e}")
        finally:
            await db.close()
    
    # 定义后台任务包装函数，确保异常被捕获
    # 同时支持任务追踪和取消
    task_key = None  # 用于追踪任务的key
    
    async def _generate_with_error_handling():
        nonlocal task_key
        try:
            result = await StoryboardService.generate_section_storyboards(
                novel_id=request.novel_id,
                template_id=request.template_id,
                llm_config_id=request.llm_config_id,
                scene_content=request.scene_content,
                scene_title=request.scene_title,
                section_number=request.section_number,
                script_id=request.script_id,
                scene_index=request.scene_index,
                style_template_id=request.style_template_id,
                inherit_prev_state=request.inherit_prev_state,
                cross_chapter_inherit=request.cross_chapter_inherit,
                with_character_state=request.with_character_state,
            )
            logger.info(f"[storyboard-api] 场景 {request.scene_index} 分镜生成完成: success={result.get('success')}, count={result.get('count')}")
            return result
        except asyncio.CancelledError:
            logger.info(f"[storyboard-api] 场景 {request.scene_index} 分镜生成任务被取消")
            raise
        except Exception as e:
            logger.error(f"[storyboard-api] 场景 {request.scene_index} 分镜生成任务异常: {type(e).__name__}: {str(e)}")
            import traceback
            traceback.print_exc()
            raise
        finally:
            # 任务完成后注销追踪
            if task_key:
                unregister_generation_task(task_key)
    
    # 启动后台任务并注册到追踪字典
    task = asyncio.create_task(_generate_with_error_handling())
    task_key = register_generation_task(
        novel_id=request.novel_id,
        script_id=request.script_id,
        scene_index=request.scene_index,
        task=task
    )
    logger.info(f"[storyboard-api] 已启动后台生成任务: task_key={task_key}, 当前活跃任务数={get_running_task_count()}")
    
    return {
        "status": "started",
        "script_id": request.script_id,
        "scene_index": request.scene_index,
        "novel_id": request.novel_id,
        "task_key": task_key
    }


@router.post("/cancel-generation")
async def cancel_generation(novel_id: int, script_id: Optional[int] = None):
    """中止正在运行的分镜生成任务(批量串行卡住时供前端中断)。

    v3.61.223: 串行批量某节 LLM 卡住(如中转 ReadTimeout 长达 10+ 分钟)时,
    前端无法等待。此接口取消该小说/章节正在跑的后台生成 asyncio 任务,
    让卡住的镜立即结束,前端轮询随即收到中止信号。
    """
    cancelled = await cancel_generation_tasks(novel_id, script_id)
    logger.info(f"[storyboard-api] /cancel-generation novel_id={novel_id} script_id={script_id} 取消 {cancelled} 个任务")
    return {"cancelled": cancelled}


@router.get("/generation-status")
async def get_generation_status(novel_id: int, script_id: int):
    """查询分镜生成状态（通过日志来源字段）
    
    增强功能：对于 status='success' 的日志，自动检查并恢复缺失的分镜数据。
    安全保护：清空分镜操作会同时删除 llm_logs，因此不会从已清空的旧日志中恢复。
    """
    db = await get_db()
    try:
        cursor = await db.execute(
            """SELECT id, status, source_scene_index, error_message, created_at
               FROM llm_logs 
               WHERE source_type = 'storyboard' 
                 AND source_id = ? 
                 AND novel_id = ?
               ORDER BY id DESC""",
            (script_id, novel_id)
        )
        rows = await cursor.fetchall()
    finally:
        await db.close()
    
    # 按 source_scene_index 分组，取每个场景的最新状态
    scene_status = {}
    for row in rows:
        idx = row['source_scene_index']
        if idx is not None and idx not in scene_status:
            scene_status[idx] = {
                "log_id": row['id'],
                "status": row['status'],
                "scene_index": idx,
                "error_message": row['error_message'],
                "created_at": row['created_at']
            }

    # 自动恢复：检查 success 状态的日志是否有对应的分镜数据
    # 安全说明：清空分镜时会同步删除 llm_logs，所以这里只会恢复「生成成功但分镜数据意外丢失」的情况
    recovered_scenes = []
    for idx, status_info in scene_status.items():
        if status_info['status'] == 'success':
            # 检查该场景是否有分镜数据
            db = await get_db()
            try:
                cursor = await db.execute(
                    "SELECT COUNT(*) as count FROM storyboards WHERE novel_id = ? AND script_id = ? AND scene_index = ?",
                    (novel_id, script_id, idx)
                )
                row = await cursor.fetchone()
                storyboard_count = row['count'] if row else 0
            finally:
                await db.close()

            if storyboard_count == 0:
                # 没有分镜数据，尝试从日志恢复
                logger.info(f"[storyboard-api] 自动恢复: 场景 {idx} 日志状态为success但无分镜数据，尝试从日志 {status_info['log_id']} 恢复")
                try:
                    recovered = await StoryboardService.recover_from_log(
                        log_id=status_info['log_id'],
                        novel_id=novel_id,
                        script_id=script_id,
                        scene_index=idx
                    )
                    if recovered:
                        status_info['recovered'] = True
                        recovered_scenes.append(idx)
                        logger.info(f"[storyboard-api] 自动恢复成功: 场景 {idx}")
                except Exception as e:
                    logger.error(f"[storyboard-api] 自动恢复失败: 场景 {idx}, 错误: {e}")

    return {
        "novel_id": novel_id,
        "script_id": script_id,
        "scenes": list(scene_status.values()),
        "recovered_scenes": recovered_scenes
    }


# 注意：精确路径路由（/grouped, /export, /import）必须放在通配路由 /novel/{novel_id} 之前
# 否则 FastAPI 会把 /novel/123/export 中的 "export" 当作下一段路径处理

@router.get("/novel/{novel_id}/grouped")
async def get_storyboards_grouped(novel_id: int, script_id: int = None):
    """获取某小说的所有分镜，按小节分组"""
    result = await StoryboardService.get_storyboards_grouped(novel_id, script_id)
    return result

@router.get("/novel/{novel_id}/export")
async def export_storyboards_excel(novel_id: int, script_id: int = None, ids: Optional[str] = None):
    """
    导出分镜到 Excel 文件
    
    Excel 列：分镜ID、章节、小节、场景序号、镜头描述、画面内容、风格提示词、人物、场景、道具
    限制：单次最多导出200条
    
    参数：
    - novel_id: 小说ID
    - script_id: 可选，指定章节ID则只导出该章节的分镜
    - ids: 可选，逗号分隔的分镜ID列表，指定则只导出这些分镜
    """
    try:
        # 解析 ids 参数
        selected_ids = None
        if ids:
            try:
                selected_ids = [int(id.strip()) for id in ids.split(",") if id.strip()]
            except ValueError:
                raise HTTPException(status_code=400, detail="ids 参数格式错误，应为逗号分隔的数字")
        
        # 获取分镜数据
        storyboards = await StoryboardService.get_storyboards(novel_id, script_id)
        
        # 如果指定了 ids，过滤出指定的分镜
        if selected_ids:
            id_set = set(selected_ids)
            storyboards = [sb for sb in storyboards if sb.get("id") in id_set]
            if not storyboards:
                raise HTTPException(status_code=404, detail="指定的分镜ID不存在或无数据")
        
        if not storyboards:
            raise HTTPException(status_code=404, detail="没有可导出的分镜数据")
        
        # 数量限制检查
        if len(storyboards) > MAX_EXPORT_COUNT:
            raise HTTPException(
                status_code=400, 
                detail=f"导出数据超过{MAX_EXPORT_COUNT}条限制（当前{len(storyboards)}条），请使用 script_id 参数分批导出"
            )
        
        # 获取剧本章节信息映射
        db = await get_db()
        script_chapter_map = {}
        try:
            cursor = await db.execute(
                """
                SELECT s.id, c.title as chapter_title 
                FROM scripts s
                LEFT JOIN chapters c ON s.chapter_id = c.id
                WHERE s.novel_id = ?
                """,
                (novel_id,)
            )
            rows = await cursor.fetchall()
            for row in rows:
                script_chapter_map[row["id"]] = row["chapter_title"] or "未知章节"
        finally:
            await db.close()
        
        # 创建 Excel 工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "分镜"
        
        # 添加表头
        headers = ["分镜ID", "章节", "小节", "场景序号", "画面内容"]
        ws.append(headers)
        
        # 设置表头样式
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)
        
        # 添加数据
        for sb in storyboards:
            chapter_title = script_chapter_map.get(sb.get("script_id"), "未知章节")
            
            ws.append([
                sb.get("id"),
                chapter_title,
                sb.get("section_number", 1),
                sb.get("scene_number", 1),
                sb.get("prompt", "") or sb.get("description", ""),
            ])
        
        # 设置列宽
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 80
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 文件大小检查
        file_size = output.getbuffer().nbytes
        if file_size > MAX_FILE_SIZE:
            raise HTTPException(
                status_code=400,
                detail=f"生成的文件过大（{file_size // 1024 // 1024}MB），超过50MB限制，请减少导出数量"
            )
        
        filename = f"storyboards_novel_{novel_id}"
        if script_id:
            filename += f"_script_{script_id}"
        filename += ".xlsx"
        
        # 使用 RFC 5987 编码文件名，支持中文
        encoded_filename = quote(filename)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


@router.post("/novel/{novel_id}/import")
async def import_storyboards_excel(novel_id: int, file: UploadFile = File(...)):
    """
    从 Excel 文件导入分镜
    
    按"分镜ID"列匹配，批量更新其他列
    限制：文件大小不超过50MB，最多处理500行
    返回：{updated_count, not_found_ids}
    """
    try:
        # 验证文件类型
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="请上传 Excel 文件 (.xlsx 或 .xls)")
        
        # 读取文件内容并检查大小
        contents = await file.read()
        file_size = len(contents)
        if file_size > MAX_FILE_SIZE:
            raise HTTPException(
                status_code=400, 
                detail=f"文件大小超过限制（{file_size // 1024 // 1024}MB > 50MB），请减少数据量"
            )
        
        wb = load_workbook(io.BytesIO(contents))
        ws = wb.active
        
        # 检查行数限制
        row_count = ws.max_row - 1  # 减去表头行
        if row_count > MAX_IMPORT_ROWS:
            raise HTTPException(
                status_code=400,
                detail=f"数据行数超过限制（{row_count}行 > {MAX_IMPORT_ROWS}行），请分批导入"
            )
        
        # 读取表头确定列位置
        headers = [cell.value for cell in ws[1] if cell.value]
        
        # 查找必需列
        try:
            id_col = headers.index("分镜ID")
        except ValueError:
            raise HTTPException(
                status_code=400, 
                detail="Excel 文件缺少必需列：分镜ID"
            )
        
        # 可选更新列的索引（可能不存在）
        optional_cols = {
            "画面内容": "prompt"
        }
        
        col_mapping = {}
        for col_name, field_name in optional_cols.items():
            try:
                col_mapping[field_name] = headers.index(col_name)
            except ValueError:
                pass
        
        if not col_mapping:
            raise HTTPException(
                status_code=400,
                detail="Excel 文件没有可更新的列（画面内容）"
            )
        
        # 解析并更新数据
        updated_count = 0
        not_found_ids = []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) <= id_col:
                continue
                
            storyboard_id = row[id_col]
            
            # 跳过空数据
            if not storyboard_id:
                continue
            
            # 尝试转换为整数
            try:
                storyboard_id = int(storyboard_id)
            except (ValueError, TypeError):
                continue
            
            # 构建更新数据
            update_data = {}
            for field_name, col_idx in col_mapping.items():
                if col_idx < len(row) and row[col_idx] is not None:
                    value = row[col_idx]
                    
                    if isinstance(value, str):
                        value = value.strip()
                    
                    update_data[field_name] = value
            
            # 画面内容同时更新 description 和 prompt（保持前端一致性）
            if "prompt" in update_data:
                update_data["description"] = update_data["prompt"]
            
            if not update_data:
                continue
            
            # 更新分镜
            result = await StoryboardService.update_storyboard(storyboard_id, update_data)
            if result:
                updated_count += 1
            else:
                not_found_ids.append(storyboard_id)
        
        return {
            "updated_count": updated_count,
            "not_found_ids": not_found_ids,
            "message": f"成功更新 {updated_count} 个分镜"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导入失败: {str(e)}")


# 适配路由放在精确路由之后
@router.get("/novel/{novel_id}")
async def get_storyboards_by_novel(novel_id: int, script_id: int = None):
    """获取某小说的所有分镜"""
    storyboards = await StoryboardService.get_storyboards(novel_id, script_id)
    return {
        "novel_id": novel_id,
        "script_id": script_id,
        "storyboards": storyboards,
        "total": len(storyboards)
    }


@router.put("/{storyboard_id}/video-status")
async def update_video_status(storyboard_id: int, video_status: str = Query(...), video_url: str = Query(None)):
    """更新分镜的视频生成状态"""
    success = await StoryboardService.update_video_status(storyboard_id, video_status, video_url)
    if success:
        return {"success": True, "message": "视频状态已更新"}
    raise HTTPException(status_code=500, detail="更新视频状态失败")


@router.put("/novel/{novel_id}/style-prompt")
async def update_storyboards_style_prompt(
    novel_id: int,
    style_prompt: str = Query(..., description="风格提示词内容"),
    script_id: int = Query(None, description="可选：指定剧本ID")
):
    """批量更新小说的所有分镜的风格提示词"""
    try:
        result = await StoryboardService.update_style_prompt_batch(novel_id, style_prompt, script_id)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"批量更新风格提示词失败: {str(e)}")


@router.post("/{storyboard_id}/extract-end-state")
async def extract_storyboard_end_state(storyboard_id: int, llm_config_id: int = Query(...)):
    """手动触发某个分镜小节的 end_state 提取(失败/空时给用户一个重试入口)。

    优先级:
      1. prompt 里有 `📎 本节结尾状态:` block → 直接解析,不调 LLM(0 风险 0 成本)
      2. 没有 → 走 LLM 提取(兜底)

    历史 bug:不管 prompt 里有没有 📎 块,都丢给 LLM,LLM 容易被 prompt 顶部的
    「场景起始状态」「渐进变化」误导,提出别节内容(2026-04 实测案例)。
    """
    from services.state_extractor_service import extract_end_state
    import json as _json
    import re as _re
    db = await get_db()
    try:
        cursor = await db.execute(
            "SELECT id, description, prompt FROM storyboards WHERE id = ?",
            (storyboard_id,)
        )
        row = await cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="分镜不存在")
        text = row["description"] or row["prompt"] or ""
    finally:
        await db.close()

    # ── 优先级 1: 解析 prompt 里的 📎/🔗 本节结尾状态 块 ──
    # 兼容两种 emoji 前缀(老模板用 🔗,新模板用 📎),以及"结尾状态结束"闭合 / 开放式两种格式
    es_block_re = _re.compile(
        r'(?:📎|🔗)\s*本节结尾状态[:：]?\s*\n'
        r'(.*?)'
        r'(?:(?:📎|🔗)\s*结尾状态结束|\n\s*(?:📏|🎬|📎|🔗)|\Z)',
        _re.DOTALL,
    )
    m = es_block_re.search(text)
    parsed_from_prompt = None
    if m:
        parsed_from_prompt = {}
        for line in m.group(1).split('\n'):
            line = line.strip()
            if not line or '=' not in line:
                continue
            name, val = line.split('=', 1)
            name = name.strip().rstrip(':：').strip()
            val = val.strip()
            if name and val:
                parsed_from_prompt[name] = val
        if not parsed_from_prompt:
            parsed_from_prompt = None

    if parsed_from_prompt:
        # 直接落库,不走 LLM
        db = await get_db()
        try:
            await db.execute(
                "UPDATE storyboards SET end_state=? WHERE id=?",
                (_json.dumps(parsed_from_prompt, ensure_ascii=False), storyboard_id)
            )
            await db.commit()
        finally:
            await db.close()
        return {"success": True, "end_state": parsed_from_prompt, "source": "prompt_block"}

    # ── 优先级 2: 走 LLM 兜底 ──
    # 把整段 text 丢给 LLM,提示词里已有强化指引(见 state_extractor_service)
    shots = [{"shot_number": 1, "description": text, "dialogue": ""}]
    es = await extract_end_state(shots=shots, llm_config_id=llm_config_id)
    if not es:
        return {"success": False, "message": "提取失败,请稍后再试或手动编辑状态"}
    db = await get_db()
    try:
        await db.execute(
            "UPDATE storyboards SET end_state=? WHERE id=?",
            (_json.dumps(es, ensure_ascii=False), storyboard_id)
        )
        await db.commit()
    finally:
        await db.close()
    return {"success": True, "end_state": es, "source": "llm"}


@router.post("/{storyboard_id}/extract-start-state")
async def extract_storyboard_start_state(storyboard_id: int):
    """手动触发某个分镜小节的 section_start_state 提取。
    两步:
    1. 优先从上节 end_state 累加过滤(主时间线继承)
    2. 兜底从本节 description 的"场景起始状态:"块抽取
    """
    import json as _json
    import re as _re

    db = await get_db()
    try:
        cursor = await db.execute(
            "SELECT id, novel_id, script_id, scene_index, section_number, sort_order, "
            "description, prompt, characters, scene_type "
            "FROM storyboards WHERE id = ?",
            (storyboard_id,)
        )
        row = await cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="分镜不存在")

        text = row["description"] or row["prompt"] or ""
        try:
            active_chars = _json.loads(row["characters"] or "[]")
        except Exception:
            active_chars = []
        section_scene_type = (row["scene_type"] or "normal").lower()
        section_start = {}

        # 1) 只有 normal 节继承主时间线累积状态
        if section_scene_type == "normal" and row["novel_id"] is not None and row["sort_order"] is not None:
            accumulated = {}
            async with db.execute(
                # v3.59.78:script_id 改严格 IS 匹配,避免跨章污染(详见 storyboard_service.py:3320 注释)
                "SELECT end_state, scene_type FROM storyboards "
                "WHERE novel_id=? AND script_id IS ? "
                "AND sort_order<? AND end_state IS NOT NULL AND end_state!='' AND end_state!='{}' "
                "ORDER BY sort_order ASC",
                (row["novel_id"], row["script_id"], row["sort_order"])
            ) as cur2:
                async for prev in cur2:
                    pst = (prev["scene_type"] or "normal").lower() if "scene_type" in prev.keys() else "normal"
                    if pst != "normal":
                        continue
                    try:
                        es = _json.loads(prev["end_state"])
                        if isinstance(es, dict):
                            accumulated.update({k: str(v) for k, v in es.items() if v})
                    except Exception:
                        pass

            # v3.61.81: 跨场景检测 — 查上一节 scene_index,跟当前 scene_index 比,不同则只继承伤势
            prev_scene_index = None
            curr_scene_index = row["scene_index"] if "scene_index" in row.keys() else None
            if curr_scene_index is not None:
                async with db.execute(
                    "SELECT scene_index FROM storyboards "
                    "WHERE novel_id=? AND script_id IS ? AND sort_order<? AND scene_index IS NOT NULL "
                    "ORDER BY sort_order DESC LIMIT 1",
                    (row["novel_id"], row["script_id"], row["sort_order"])
                ) as cur3:
                    prow = await cur3.fetchone()
                    if prow:
                        prev_scene_index = prow["scene_index"]

            is_scene_change = (
                prev_scene_index is not None
                and curr_scene_index is not None
                and curr_scene_index != prev_scene_index
            )

            if is_scene_change:
                # 主线跨场景:只继承伤势,其他字段填占位,等下面 description 抽取阶段补
                section_start = {}
                for c in active_chars:
                    if c in accumulated:
                        inj_m = _re.search(r'伤势\[([^\]]+)\]', accumulated[c])
                        injury = inj_m.group(1) if inj_m else None
                        if injury and injury not in ('无伤', '无', ''):
                            section_start[c] = (
                                f"姿态[待 description 提取] · "
                                f"伤势[{injury}] · "
                                f"持有道具[待 description 提取] · "
                                f"情绪[待 description 提取] · "
                                f"朝向关系[待 description 提取]"
                            )
                # 在闭包外标记一下,供下面 description 抽取阶段判断
                _is_scene_change_flag = True
            else:
                # 同场景:走原逻辑(全盘继承,防 LLM 演绎)
                section_start = {c: accumulated[c] for c in active_chars if c in accumulated}
                _is_scene_change_flag = False
        else:
            _is_scene_change_flag = False

        # 2) 兜底 / 跨场景补足:从 description 的"场景起始状态:"块抽取
        # - 同场景:只补充 section_start 里没有的角色(首次出场)
        # - 跨场景:用 description 覆盖占位姿态/朝向/情绪/道具,但保留累积的真实伤势
        if not section_start or len(section_start) < len(active_chars) or _is_scene_change_flag:
            m = _re.search(
                r'场景起始状态\s*[:：]?\s*\n'
                r'((?:[ \t]+[^\n]+\n?)+)',
                text
            )
            if m:
                extracted = {}
                for line in m.group(1).split('\n'):
                    line = line.strip()
                    if not line or '=' not in line:
                        continue
                    name, val = line.split('=', 1)
                    name = name.strip().rstrip(':：').strip()
                    val = val.strip()
                    if name and val:
                        extracted[name] = val
                # 只取激活角色白名单内的
                filtered = {c: extracted[c] for c in active_chars if c in extracted}
                for k, v in filtered.items():
                    if k not in section_start:
                        # 不在 section_start 里(首次出场 / 跨场景无伤角色)→ 直接用 description
                        section_start[k] = v
                    elif _is_scene_change_flag:
                        # 跨场景:section_start[k] 是占位结构,用 description 覆盖,但保留累积的真实伤势
                        existing_inj_m = _re.search(r'伤势\[([^\]]+)\]', section_start[k])
                        existing_injury = existing_inj_m.group(1) if existing_inj_m else None
                        if existing_injury and existing_injury not in ('待 description 提取', '无伤', '无', ''):
                            section_start[k] = _re.sub(
                                r'伤势\[[^\]]*\]',
                                f'伤势[{existing_injury}]',
                                v
                            )
                        else:
                            section_start[k] = v

        # 写库
        await db.execute(
            "UPDATE storyboards SET section_start_state=? WHERE id=?",
            (_json.dumps(section_start, ensure_ascii=False), storyboard_id)
        )
        await db.commit()
    finally:
        await db.close()

    return {
        "success": True,
        "section_start_state": section_start,
        "count": len(section_start),
    }


@router.post("/regenerate-single-section")
async def regenerate_single_section(request: RegenerateSingleSectionRequest):
    """重新生成单个小节的分镜"""
    logger.info(f"[storyboard-api] /regenerate-single-section 被调用: novel_id={request.novel_id}, "
          f"storyboard_id={request.storyboard_id}, section_number={request.section_number}, "
          f"scene_title={request.scene_title}, scene_index={request.scene_index}")
    try:
        result = await StoryboardService.regenerate_single_section(
            novel_id=request.novel_id,
            template_id=request.template_id,
            llm_config_id=request.llm_config_id,
            scene_content=request.scene_content,
            scene_title=request.scene_title,
            section_number=request.section_number,
            storyboard_id=request.storyboard_id,
            script_id=request.script_id,
            style_template_id=request.style_template_id,
            inherit_prev_state=request.inherit_prev_state,
            cross_chapter_inherit=request.cross_chapter_inherit,
            with_character_state=request.with_character_state,
        )

        logger.info(f"[storyboard-api] /regenerate-single-section 服务返回: success={result.get('success')}, "
              f"message={result.get('message')}")

        if not result["success"]:
            raise HTTPException(status_code=400, detail=result["message"])

        return result

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[storyboard-api] /regenerate-single-section 发生异常: {type(e).__name__}: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"重新生成小节分镜失败: {str(e)}")


# 以下路由使用 /{storyboard_id} 通配符，必须放在所有 /novel/xxx 路由之后

@router.get("/{storyboard_id}")
async def get_storyboard(storyboard_id: int):
    """获取单个分镜"""
    storyboard = await StoryboardService.get_storyboard(storyboard_id)
    if not storyboard:
        raise HTTPException(status_code=404, detail="分镜不存在")
    return storyboard


@router.put("/{storyboard_id}")
async def update_storyboard(storyboard_id: int, data: StoryboardUpdate):
    """更新分镜"""
    update_data = data.model_dump(exclude_unset=True)
    if (
        "section_start_state" not in update_data
        and ("description" in update_data or "prompt" in update_data)
    ):
        state_text = update_data.get("description") or update_data.get("prompt") or ""
        parsed_start_state = _parse_section_start_state_from_text(state_text)
        if parsed_start_state:
            update_data["section_start_state"] = parsed_start_state
            logger.info(
                "[storyboard-api] 同步用户编辑的 section_start_state: "
                f"storyboard_id={storyboard_id}, count={len(parsed_start_state)}"
            )
    storyboard = await StoryboardService.update_storyboard(storyboard_id, update_data)
    if not storyboard:
        raise HTTPException(status_code=404, detail="分镜不存在")
    return storyboard


# ==================== 敏感词扫描/清洗 ====================

class SensitiveScanRequest(BaseModel):
    text: str


@router.post("/sensitive/scan")
async def scan_sensitive_words(req: SensitiveScanRequest):
    """扫描文本中的敏感词,返回命中列表 + 清洗后文本(不落盘)"""
    result = await sensitive_word_service.scan_text(req.text or "")
    return result


class SensitiveApplyRequest(BaseModel):
    storyboard_id: int
    cleaned_text: str  # 前端已确认的清洗结果


@router.post("/sensitive/apply")
async def apply_cleaned_prompt(req: SensitiveApplyRequest):
    """把用户确认后的清洗文本写回分镜 prompt 字段"""
    db = await get_db()
    try:
        cur = await db.execute("SELECT id FROM storyboards WHERE id = ?", (req.storyboard_id,))
        if not await cur.fetchone():
            raise HTTPException(status_code=404, detail="分镜不存在")
        await db.execute(
            "UPDATE storyboards SET prompt = ?, description = ? WHERE id = ?",
            (req.cleaned_text, req.cleaned_text, req.storyboard_id)
        )
        await db.commit()
    finally:
        await db.close()
    return {"success": True, "storyboard_id": req.storyboard_id}


@router.delete("/{storyboard_id}")
async def delete_storyboard(storyboard_id: int):
    """删除分镜"""
    storyboard = await StoryboardService.get_storyboard(storyboard_id)
    if not storyboard:
        raise HTTPException(status_code=404, detail="分镜不存在")
    if not await NovelService.get_by_id(storyboard.get("novel_id")):
        raise HTTPException(status_code=404, detail="分镜不存在或不属于当前账号")
    success = await StoryboardService.delete_storyboard(storyboard_id)
    if not success:
        raise HTTPException(status_code=404, detail="分镜不存在")
    return {"message": "删除成功"}


@router.delete("/novel/{novel_id}/all")
async def delete_storyboards_by_novel(novel_id: int, script_id: int = None):
    """清空小说分镜，如果指定了script_id则只清空该章节的分镜"""
    count = await StoryboardService.delete_storyboards_by_novel(novel_id, script_id)
    return {
        "message": f"已删除 {count} 个分镜",
        "deleted_count": count
    }


@router.put("/novel/{novel_id}/reorder")
async def reorder_storyboards(novel_id: int, storyboard_ids: List[int]):
    """重新排序分镜"""
    success = await StoryboardService.reorder_storyboards(novel_id, storyboard_ids)
    if not success:
        raise HTTPException(status_code=400, detail="排序失败")
    return {"message": "排序成功"}


@router.post("/fix-empty-scenes")
async def fix_empty_scenes(novel_id: int = Query(..., description="小说ID"), script_id: Optional[int] = Query(None, description="剧本ID")):
    """修复分镜中场景字段为空的记录，根据剧本场景拆分结果填充场景名"""
    logger.info(f"[storyboard-api] /fix-empty-scenes 被调用: novel_id={novel_id}, script_id={script_id}")
    try:
        result = await StoryboardService.fix_empty_scene_fields(novel_id, script_id)
        return result
    except Exception as e:
        logger.error(f"[storyboard-api] 修复空场景失败: {e}")
        raise HTTPException(status_code=500, detail=f"修复失败: {str(e)}")
