from fastapi import APIRouter, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from typing import Optional
from database.db import get_db
from models.scripts import (
    ScriptConvertRequest,
    ScriptResponse,
    ScriptListResponse,
    ScriptUpdateRequest,
    ScriptConvertResponse,
    SingleScriptConvertRequest,
    ScriptConvertResult,
    OfficialScriptResultRequest
)
from services.script_service import ScriptService
from services.log_service import check_novel_running
import io
from openpyxl import Workbook, load_workbook
from urllib.parse import quote
from utils.timezone import now_beijing_str

router = APIRouter(prefix="/api/scripts", tags=["scripts"])


@router.post("/convert", response_model=ScriptConvertResponse)
async def convert_scripts(request: ScriptConvertRequest):
    """
    启动剧本转换（批量）
    
    - 如果指定了 chapter_ids，则只转换这些章节
    - 否则转换小说的所有章节
    """
    # 冲突检查
    db = await get_db()
    try:
        await check_novel_running(request.novel_id, db)
    finally:
        await db.close()

    try:
        result = await ScriptService.convert_all_chapters(
            novel_id=request.novel_id,
            template_id=request.template_id,
            llm_config_id=request.llm_config_id,
            chapter_ids=request.chapter_ids
        )
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"转换失败: {str(e)}")


@router.post("/convert-single", response_model=ScriptConvertResult)
async def convert_single_script(request: SingleScriptConvertRequest):
    """
    单章节剧本转换
    
    用于前端逐章节转换，实时显示进度
    """
    # 冲突检查
    db = await get_db()
    try:
        await check_novel_running(request.novel_id, db)
    finally:
        await db.close()

    try:
        result = await ScriptService.convert_chapter(
            novel_id=request.novel_id,
            chapter_id=request.chapter_id,
            template_id=request.template_id,
            llm_config_id=request.llm_config_id
        )
        return {
            "chapter_id": request.chapter_id,
            "chapter_title": result["chapter_title"],
            "success": result["success"],
            "script_id": result.get("script_id"),
            "message": result["message"]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"转换失败: {str(e)}")


@router.post("/official-result", response_model=ScriptResponse)
async def save_official_script_result(request: OfficialScriptResultRequest):
    """保存已由 anyq.site 官方语言算力生成的正文，不在本地再次调用模型。"""
    content = str(request.content or "").strip()
    if not content:
        raise HTTPException(status_code=400, detail="官方任务没有返回可保存的剧本正文")
    db = await get_db()
    try:
        chapter_cursor = await db.execute(
            "SELECT id FROM chapters WHERE id=? AND novel_id=?",
            (request.chapter_id, request.novel_id)
        )
        if not await chapter_cursor.fetchone():
            raise HTTPException(status_code=404, detail="章节不存在或不属于当前小说")
        cursor = await db.execute(
            "SELECT id FROM scripts WHERE novel_id=? AND chapter_id=? ORDER BY id LIMIT 1",
            (request.novel_id, request.chapter_id)
        )
        existing = await cursor.fetchone()
        if existing:
            await db.execute(
                "UPDATE scripts SET content=?, template_id=COALESCE(?, template_id), remote_version=-1 WHERE id=?",
                (content, request.template_id, existing["id"])
            )
            script_id = existing["id"]
        else:
            cursor = await db.execute(
                "INSERT INTO scripts (novel_id, chapter_id, content, template_id, remote_version, created_at) VALUES (?, ?, ?, ?, -1, ?)",
                (request.novel_id, request.chapter_id, content, request.template_id, now_beijing_str())
            )
            script_id = cursor.lastrowid
        await db.commit()
    finally:
        await db.close()
    result = await ScriptService.get_script(script_id)
    if not result:
        raise HTTPException(status_code=500, detail="剧本保存后读取失败")
    return result


# 注意：精确路径路由（/export, /import, /all）必须放在通配路由 /novel/{novel_id} 之前
# 否则 FastAPI 会把 /novel/123/export 中的 "export" 当作下一段路径处理

@router.get("/novel/{novel_id}/export")
async def export_scripts_excel(novel_id: int, chapter_ids: Optional[str] = None):
    """
    导出剧本到 Excel 文件
    
    Excel 列：剧本ID、章节ID、章节标题、剧本内容
    
    参数：
    - novel_id: 小说ID
    - chapter_ids: 可选，逗号分隔的章节ID列表，指定则只导出这些章节的剧本
    """
    try:
        db = await get_db()
        try:
            # 构建查询
            if chapter_ids:
                # 解析章节ID列表
                try:
                    id_list = [int(id.strip()) for id in chapter_ids.split(",") if id.strip()]
                except ValueError:
                    raise HTTPException(status_code=400, detail="chapter_ids 参数格式错误，应为逗号分隔的数字")

                if not id_list:
                    raise HTTPException(status_code=400, detail="chapter_ids 不能为空")
                
                placeholders = ",".join(["?" for _ in id_list])
                cursor = await db.execute(
                    f"""
                    SELECT s.id, s.chapter_id, c.title as chapter_title, s.content
                    FROM scripts s
                    LEFT JOIN chapters c ON s.chapter_id = c.id
                    WHERE s.novel_id = ? AND s.chapter_id IN ({placeholders})
                    ORDER BY c.sort_order, s.chapter_id
                    """,
                    (novel_id, *id_list)
                )
            else:
                cursor = await db.execute(
                    """
                    SELECT s.id, s.chapter_id, c.title as chapter_title, s.content
                    FROM scripts s
                    LEFT JOIN chapters c ON s.chapter_id = c.id
                    WHERE s.novel_id = ?
                    ORDER BY c.sort_order, s.chapter_id
                    """,
                    (novel_id,)
                )
            
            rows = await cursor.fetchall()
            
            if not rows:
                raise HTTPException(status_code=404, detail="没有可导出的剧本数据")
            
            # 创建 Excel 工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "剧本"
            
            # 添加表头
            headers = ["剧本ID", "章节ID", "章节标题", "剧本内容"]
            ws.append(headers)
            
            # 设置表头样式
            for cell in ws[1]:
                cell.font = cell.font.copy(bold=True)
            
            # 添加数据
            for row in rows:
                ws.append([
                    row["id"],
                    row["chapter_id"],
                    row["chapter_title"] or "未知章节",
                    row["content"] or ""
                ])
            
            # 设置列宽
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 100
            
            # 保存到内存
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f"scripts_novel_{novel_id}.xlsx"
            encoded_filename = quote(filename)
            
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
                }
            )
        finally:
            await db.close()
            
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


@router.get("/novel/{novel_id}/export-txt")
async def export_scripts_txt(novel_id: int, chapter_ids: Optional[str] = None):
    """
    导出剧本到 txt 文件（场景标头 + 台词原文）

    每集格式：章节标题 + 空行 + 剧本正文；集与集之间用分隔线隔开。

    参数：
    - novel_id: 小说ID
    - chapter_ids: 可选，逗号分隔的章节ID列表，指定则只导出这些章节的剧本
    """
    try:
        db = await get_db()
        try:
            # 取小说名作为文件名
            novel_cursor = await db.execute(
                "SELECT name FROM novels WHERE id = ?",
                (novel_id,)
            )
            novel_row = await novel_cursor.fetchone()
            if not novel_row:
                raise HTTPException(status_code=404, detail="小说不存在")
            novel_name = novel_row["name"] or f"novel_{novel_id}"

            # 构建查询
            if chapter_ids:
                try:
                    id_list = [int(id.strip()) for id in chapter_ids.split(",") if id.strip()]
                except ValueError:
                    raise HTTPException(status_code=400, detail="chapter_ids 参数格式错误，应为逗号分隔的数字")

                if not id_list:
                    raise HTTPException(status_code=400, detail="chapter_ids 不能为空")

                placeholders = ",".join(["?" for _ in id_list])
                cursor = await db.execute(
                    f"""
                    SELECT s.id, s.chapter_id, c.title as chapter_title, s.content
                    FROM scripts s
                    LEFT JOIN chapters c ON s.chapter_id = c.id
                    WHERE s.novel_id = ? AND s.chapter_id IN ({placeholders})
                    ORDER BY c.sort_order, s.chapter_id
                    """,
                    (novel_id, *id_list)
                )
            else:
                cursor = await db.execute(
                    """
                    SELECT s.id, s.chapter_id, c.title as chapter_title, s.content
                    FROM scripts s
                    LEFT JOIN chapters c ON s.chapter_id = c.id
                    WHERE s.novel_id = ?
                    ORDER BY c.sort_order, s.chapter_id
                    """,
                    (novel_id,)
                )

            rows = await cursor.fetchall()

            if not rows:
                raise HTTPException(status_code=404, detail="没有可导出的剧本数据")

            # 拼接文本：每集「章节标题\n\n正文」，集间用分隔线
            separator = "\n\n" + "=" * 30 + "\n\n"
            blocks = []
            for row in rows:
                title = row["chapter_title"] or "未知章节"
                content = row["content"] or ""
                blocks.append(f"{title}\n\n{content}")
            text = separator.join(blocks)

            output = io.BytesIO(text.encode("utf-8"))
            output.seek(0)

            encoded_filename = quote(f"{novel_name}_剧本.txt")

            return StreamingResponse(
                output,
                media_type="text/plain; charset=utf-8",
                headers={
                    "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
                }
            )
        finally:
            await db.close()

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


@router.post("/novel/{novel_id}/import")
async def import_scripts_excel(novel_id: int, file: UploadFile = File(...)):
    """
    从 Excel 文件导入剧本
    
    按"剧本ID"或"章节ID"列匹配，批量更新剧本内容
    返回：{updated_count, not_found_ids}
    """
    try:
        # 验证文件类型
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="请上传 Excel 文件 (.xlsx 或 .xls)")
        
        # 读取文件内容
        contents = await file.read()
        
        wb = load_workbook(io.BytesIO(contents))
        ws = wb.active
        
        # 读取表头确定列位置
        headers = [cell.value for cell in ws[1] if cell.value]
        
        # 查找必需列
        try:
            script_id_col = headers.index("剧本ID")
        except ValueError:
            raise HTTPException(status_code=400, detail="Excel 文件缺少必需列：剧本ID")
        
        # 查找可选列
        chapter_id_col = None
        content_col = None
        try:
            chapter_id_col = headers.index("章节ID")
        except ValueError:
            pass
        try:
            content_col = headers.index("剧本内容")
        except ValueError:
            pass
        
        if content_col is None:
            raise HTTPException(status_code=400, detail="Excel 文件缺少必需列：剧本内容")
        
        db = await get_db()
        try:
            updated_count = 0
            not_found_ids = []
            
            # 遍历数据行
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                script_id = row[script_id_col]
                content = row[content_col] if content_col is not None else None
                
                if not script_id:
                    continue
                
                # 检查剧本是否存在且属于该小说
                cursor = await db.execute(
                    "SELECT id FROM scripts WHERE id = ? AND novel_id = ?",
                    (script_id, novel_id)
                )
                existing = await cursor.fetchone()
                
                if not existing:
                    not_found_ids.append(script_id)
                    continue
                
                # 更新剧本内容
                # v3.61.142:批量导入也算"本地编辑",打 dirty 标记防被短剧同步覆盖
                if content is not None:
                    await db.execute(
                        "UPDATE scripts SET content = ?, remote_version = -1 WHERE id = ?",
                        (content, script_id)
                    )
                    updated_count += 1
            
            await db.commit()
            
            return {
                "updated_count": updated_count,
                "not_found_ids": not_found_ids
            }
        finally:
            await db.close()
            
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导入失败: {str(e)}")


@router.delete("/novel/{novel_id}/all")
async def delete_scripts_by_novel(novel_id: int):
    """删除小说的所有剧本"""
    count = await ScriptService.delete_scripts_by_novel(novel_id)
    return {"message": f"已删除 {count} 个剧本"}


@router.get("/novel/{novel_id}", response_model=ScriptListResponse)
async def get_scripts_by_novel(novel_id: int):
    """获取某小说的所有剧本"""
    try:
        result = await ScriptService.get_scripts(novel_id)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取剧本列表失败: {str(e)}")


@router.get("/{script_id}", response_model=ScriptResponse)
async def get_script(script_id: int):
    """获取单个剧本"""
    script = await ScriptService.get_script(script_id)
    if not script:
        raise HTTPException(status_code=404, detail="剧本不存在")
    return script


@router.get("/chapter/{chapter_id}", response_model=ScriptResponse)
async def get_script_by_chapter(chapter_id: int):
    """根据章节ID获取剧本"""
    script = await ScriptService.get_script_by_chapter(chapter_id)
    if not script:
        raise HTTPException(status_code=404, detail="该章节尚未转换为剧本")
    return script


@router.put("/{script_id}", response_model=ScriptResponse)
async def update_script(script_id: int, request: ScriptUpdateRequest):
    """更新剧本内容（支持手动编辑）"""
    script = await ScriptService.update_script(script_id, request.content)
    if not script:
        raise HTTPException(status_code=404, detail="剧本不存在")
    return script


@router.put("/{script_id}/scene-meta")
async def update_script_scene_meta(script_id: int, request: dict):
    """更新剧本的场景元数据(scene_meta),用于在分镜生成前预先标注场景类型

    Body: {"scene_index": 0, "scene_type": "flashback"}
    或批量: {"meta": {"0": {"scene_type": "flashback"}, "2": {"scene_type": "dream"}}}
    """
    import json
    from database.db import get_db
    db = await get_db()
    try:
        async with db.execute("SELECT scene_meta FROM scripts WHERE id=?", (script_id,)) as cur:
            row = await cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="剧本不存在")
            try:
                meta = json.loads(row["scene_meta"] or "{}")
            except Exception:
                meta = {}

        if "meta" in request and isinstance(request["meta"], dict):
            # 批量覆盖
            for k, v in request["meta"].items():
                meta[str(k)] = v
        elif "scene_index" in request and "scene_type" in request:
            # 单个更新
            idx = str(request["scene_index"])
            meta[idx] = meta.get(idx, {})
            meta[idx]["scene_type"] = request["scene_type"]
        else:
            raise HTTPException(status_code=400, detail="参数无效,需要 {scene_index, scene_type} 或 {meta}")

        await db.execute(
            "UPDATE scripts SET scene_meta=? WHERE id=?",
            (json.dumps(meta, ensure_ascii=False), script_id)
        )
        await db.commit()
        return {"success": True, "scene_meta": meta}
    finally:
        await db.close()


@router.get("/{script_id}/scene-meta")
async def get_script_scene_meta(script_id: int):
    """获取剧本的 scene_meta"""
    import json
    from database.db import get_db
    db = await get_db()
    try:
        async with db.execute("SELECT scene_meta FROM scripts WHERE id=?", (script_id,)) as cur:
            row = await cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="剧本不存在")
            try:
                meta = json.loads(row["scene_meta"] or "{}")
            except Exception:
                meta = {}
            return {"success": True, "scene_meta": meta}
    finally:
        await db.close()


@router.get("/chapter/{chapter_id}/prev-scene")
async def get_prev_chapter_last_scene(chapter_id: int, novel_id: int = Query(...)):
    """返回当前章节的上一章剧本的最后一个场景(用于 UI 提示"上一章末场景")。
    如果上一章没剧本或没场景返回 {success:True, data:None}。
    """
    try:
        data = await ScriptService._get_prev_chapter_last_scene(novel_id, chapter_id)
        return {"success": True, "data": data}
    except Exception as e:
        return {"success": False, "data": None, "message": str(e)}


@router.delete("/{script_id}")
async def delete_script(script_id: int):
    """删除剧本"""
    success = await ScriptService.delete_script(script_id)
    if not success:
        raise HTTPException(status_code=404, detail="剧本不存在")
    return {"message": "剧本已删除"}
